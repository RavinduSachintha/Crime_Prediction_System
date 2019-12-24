import os
from math import isnan

import pandas as pd
from bokeh.models import ColumnDataSource, TableColumn, DataTable
from openpyxl import load_workbook


class DataFile:
    filename = os.path.join(os.path.dirname(__file__), 'uploads', 'SL_CrimeDataSet.xlsx')

    def __init__(self):
        self.sheet = load_workbook(self.filename).active
        self.df = pd.read_excel(self.filename, skiprows=1)

    def get_table_column_titles(self):
        return [cell.value for cell in self.sheet[self.sheet.min_row]]

    def get_table_data(self, rows):
        table_data = []
        for i in range(self.sheet.min_row + 1, self.sheet.min_row + rows + 1):
            table_data.append([cell.value for cell in self.sheet[i]])
        return table_data

    def get_table_by_area(self, area, rows):
        lst = [row for row in self.sheet.iter_rows() if row[1].value == str(area)][:rows]
        data = dict(
            crimes=[row[2].value for row in lst],
            months=[row[4].value for row in lst],
            days=[row[5].value for row in lst],
            hours=[row[6].value for row in lst],
            penalties=[row[7].value for row in lst],
            temperature=[row[9].value for row in lst],
            pop_density=[row[10].value for row in lst],
            edu_level=[row[11].value for row in lst],
            economy=[row[12].value for row in lst],
            alcohol_drug_usage=[row[13].value for row in lst]
        )
        source = ColumnDataSource(data)
        columns = [
            TableColumn(field="crimes", title="Crime"),
            TableColumn(field="months", title="Month"),
            TableColumn(field="days", title="Date"),
            TableColumn(field="hours", title="Hour"),
            TableColumn(field="penalties", title="Penalty"),
            TableColumn(field="temperature", title="Temperature"),
            TableColumn(field="pop_density", title="Population Density"),
            TableColumn(field="edu_level", title="Education Level"),
            TableColumn(field="economy", title="Economy"),
            TableColumn(field="alcohol_drug_usage", title="Alcohol & Drug Usage"),
        ]
        return DataTable(source=source, columns=columns, sizing_mode="scale_both")

    def get_area_choices(self):
        return {str(i): idx for idx, i in enumerate(self.df['Area'].unique()) if isinstance(i, str)}

    def get_month_choices(self):
        return {str(int(i)): i for i in sorted(self.df['Month'].unique()) if not isnan(i)}

    def get_day_choices(self):
        return {str(int(i)): i for i in sorted(self.df['Day'].unique()) if not isnan(i)}

    def get_hour_choices(self):
        return {str(int(i)): i for i in sorted(self.df['Hour'].unique()) if not isnan(i)}
